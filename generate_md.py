"""生成秋招信息 Markdown 文件（适配新表结构）"""
from openpyxl import load_workbook
from datetime import datetime
import os

wb = load_workbook('/workspace/秋招投递追踪表.xlsx')
ws = wb['秋招投递追踪']
rows = list(ws.iter_rows(values_only=True))
today = datetime.now().strftime("%Y-%m-%d")

# 新表结构: 序号,记录日期,公司名称,所属行业,岗位名称,岗位类别,工作地点,学历要求,毕业时间要求,网申开始日期,网申截止日期,投递方式/链接,信息来源,是否已投递,投递日期,当前状态,备注
# 映射到旧的展示格式: 公司,岗位,类别,地点,学历,_,投递链接,开始,截止,来源,记录日期,备注,状态

def map_row(row):
    """将新表行映射到旧格式"""
    company = row[2] or ""
    job = row[4] or ""
    category = row[5] or "其他"
    location = row[6] or ""
    edu = row[7] or "硕士及以上"
    apply = row[11] or ""
    start = row[9] or ""
    end = row[10] or ""
    source = row[12] or ""
    collect_date = row[1] or today
    note = row[16] or ""
    status_val = row[15] or "待投递"
    status = "已截止" if status_val == "已截止" else "开放中"
    return (company, job, category, location, edu, "", apply, start, end, source, collect_date, note, status)

mapped_rows = [map_row(r) for r in rows[1:] if r[15] != "已截止"]  # 只展示未截止的

# 大公司名单
big_companies = {
    "工商银行", "建设银行", "农业银行", "中国银行", "交通银行", "邮储银行",
    "招商银行", "兴业银行", "浦发银行", "中信银行", "民生银行", "平安银行",
    "宁波银行", "南京银行", "杭州银行", "上海银行", "北京银行", "广发银行",
    "华夏银行", "光大银行", "浙商银行", "渤海银行",
    "中国邮政储蓄银行", "重庆银行", "广州银行", "四川银行", "温州银行",
    "中信证券", "中金公司", "华泰证券", "国泰君安", "海通证券", "申万宏源",
    "招商证券", "广发证券", "中国银河", "中信建投", "国信证券", "东方证券",
    "兴业证券", "光大证券", "平安证券", "中泰证券", "长江证券",
    "财通证券", "国投证券", "中国银河证券", "国泰海通证券", "万联证券",
    "易方达", "华夏基金", "广发基金", "南方基金", "富国基金", "嘉实基金",
    "博时基金", "招商基金", "天弘基金", "汇添富", "中庚基金", "泰康资产",
    "中粮信托", "中联基金", "拾贝投资", "平方和投资", "线性资本", "九合创投",
    "中国人寿", "中国平安", "中国人保", "中国太保", "新华保险", "泰康保险",
    "阳光保险", "中华财险", "中国大地保险", "瑞众保险",
    "腾讯", "阿里巴巴", "字节跳动", "百度", "美团", "京东", "拼多多",
    "网易", "小米", "快手", "滴滴", "小红书", "B站", "携程", "同花顺",
    "大疆", "华为", "联想", "OPPO", "vivo", "蚂蚁集团", "360",
    "宝洁", "联合利华", "欧莱雅", "雀巢", "百事", "可口可乐", "玛氏",
    "亿滋", "蒙牛", "伊利", "安踏", "李宁", "耐克", "阿迪达斯",
    "沃尔玛", "山姆", "卡地亚", "滔搏", "ALDI", "美赞臣", "益海嘉里",
    "康师傅", "Babycare", "曼伦", "欧莱雅百库",
    "普华永道", "德勤", "安永", "毕马威", "埃森哲",
    "万科", "碧桂园", "保利", "华润", "绿城", "龙湖", "招商蛇口",
    "金地", "中国铁建地产", "象屿", "中建", "中国铁建",
    "成都高新投资", "上海国际集团", "北京首都科技发展", "国新国际",
    "比亚迪", "特斯拉", "蔚来", "理想", "小鹏", "宁德时代", "上汽",
    "一汽", "吉利", "中冶南方", "中工国际", "中化商务", "中国联合工程",
    "南方财经", "人民日报", "中国建筑出版传媒", "中国邮政",
    "上汽安吉物流", "大连商品交易所", "广州碳排放权交易中心",
    "桂林理工大学", "西安铁一中教育集团", "江苏省盐业集团",
    "华润银行", "华润集团", "中粮", "中化", "中国诚通", "中国国新",
    "中国移动", "中国电信", "中国联通", "国家电网", "南方电网",
    "中国商飞", "中国航发", "凯辉集团", "策马集团",
}

def is_big(company):
    company = str(company)
    for bc in big_companies:
        if bc in company:
            return True
    return False

# 按类别分组
by_category = {}
for row in mapped_rows:
    cat = row[2] if row[2] else "其他"
    if cat not in by_category:
        by_category[cat] = []
    by_category[cat].append(row)

cat_order = ["金融", "战略", "商分", "运营", "管培", "市场", "产品", "咨询", "职能", "供应链", "传媒", "零售"]
cats_sorted = sorted(by_category.items(), key=lambda x: (cat_order.index(x[0]) if x[0] in cat_order else 99, -len(x[1])))

# 主索引
lines = []
lines.append('# 2026 秋招投递追踪')
lines.append('')
lines.append(f'> 每天早上 6:00 自动更新 | 最后更新：{today} | 共 {len(mapped_rows)} 条开放岗位')
lines.append('')
lines.append('📱 手机和网页端均可直接查看。')
lines.append('')
lines.append('## 大公司精选')
lines.append('')
lines.append(f'[大公司投递表](./docs/大公司投递表.md) — 跨行业大厂/国企/央企精选')
lines.append('')
lines.append('## 岗位分类')
lines.append('')
lines.append('| 分类 | 数量 |')
lines.append('|------|------|')
for cat, items in cats_sorted:
    safe_name = cat.replace("/", "-")
    lines.append(f'| [{cat}](./docs/{safe_name}.md) | {len(items)} 条 |')
lines.append('')
lines.append('---')
lines.append('')
lines.append('同时提供 [Excel版本](./秋招投递追踪表.xlsx)，下载后可筛选排序、标记已投递。')
lines.append('')
lines.append('*每日自动更新 · 祝秋招顺利！*')

with open('/workspace/autumn_recruitment.md', 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

# 各分类文件
os.makedirs('/workspace/docs', exist_ok=True)
for cat, items in cats_sorted:
    safe_name = cat.replace("/", "-")
    # 排序：大公司+开放中优先
    items_sorted = sorted(items, key=lambda r: (not is_big(r[0]), r[12] != "开放中"))
    
    clines = [f'# {cat}']
    clines.append('')
    open_count = sum(1 for r in items_sorted if r[12] == "开放中")
    clines.append(f'> 共 {len(items)} 条 | 开放中 {open_count} | 更新于 {today}')
    clines.append('')
    clines.append('| 公司 | 岗位 | 地点 | 截止 | 投递 | 备注 | 状态 |')
    clines.append('|------|------|------|------|------|------|------|')
    
    for row in items_sorted:
        company, job, category, location, edu, _, apply, start, end, source, collect_date, note, status = row
        status_e = '🟢' if status == '开放中' else '🔴'
        company = (company or '—')[:20]
        job = (job or '—')[:35]
        location = (location or '—')[:12]
        end = (end or '—')[:10]
        note = (note or '—')[:20]
        apply_cell = f'[投递]({apply})' if apply else '—'
        clines.append(f'| {company} | {job} | {location} | {end} | {apply_cell} | {note} | {status_e} {status} |')
    
    clines.append('')
    clines.append(f'[← 返回首页](../autumn_recruitment.md)')
    
    with open(f'/workspace/docs/{safe_name}.md', 'w', encoding='utf-8') as f:
        f.write('\n'.join(clines))

# 大公司专属表
big_items = [(r, is_big(r[0])) for r in mapped_rows if is_big(r[0])]
big_items.sort(key=lambda x: (x[0][12] != "开放中"))

big_by_cat = {}
for row, _ in big_items:
    cat = row[2] if row[2] else "其他"
    if cat not in big_by_cat:
        big_by_cat[cat] = []
    big_by_cat[cat].append(row)

big_cats_sorted = sorted(big_by_cat.items(), key=lambda x: (cat_order.index(x[0]) if x[0] in cat_order else 99, -len(x[1])))

blines = []
blines.append('# 大公司投递表')
blines.append('')
blines.append(f'> 精选大厂/国企/央企/知名企业 | 更新于 {today} | 共 {len(big_items)} 条')
blines.append('')
blines.append('跨行业精选，只要有你能做的泛商科岗位都收录了。')
blines.append('')
blines.append('[← 返回全部岗位](../autumn_recruitment.md) | [Excel版本](../秋招投递追踪表.xlsx)')
blines.append('')

for cat, items in big_cats_sorted:
    items_sorted = sorted(items, key=lambda r: (r[12] != "开放中"))
    open_count = sum(1 for r in items_sorted if r[12] == "开放中")
    blines.append(f'## {cat} ({len(items)}条)')
    blines.append('')
    blines.append(f'> 开放中 {open_count}')
    blines.append('')
    blines.append('| 公司 | 岗位 | 地点 | 截止 | 投递 | 备注 | 状态 |')
    blines.append('|------|------|------|------|------|------|------|')
    
    for row in items_sorted:
        company, job, category, location, edu, _, apply, start, end, source, collect_date, note, status = row
        status_e = '🟢' if status == '开放中' else '🔴'
        company = (company or '—')[:20]
        job = (job or '—')[:35]
        location = (location or '—')[:12]
        end = (end or '—')[:10]
        note = (note or '—')[:20]
        apply_cell = f'[投递]({apply})' if apply else '—'
        blines.append(f'| {company} | {job} | {location} | {end} | {apply_cell} | {note} | {status_e} {status} |')
    blines.append('')

blines.append('---')
blines.append('*每日自动更新 · 祝秋招顺利！*')

with open('/workspace/docs/大公司投递表.md', 'w', encoding='utf-8') as f:
    f.write('\n'.join(blines))

# 导出 JSON 供前端页面使用
import json
json_data = []
for row in mapped_rows:
    company, job, category, location, edu, _, apply, start, end, source, collect_date, note, status = row
    json_data.append({
        "company": company, "position": job, "category": category,
        "location": location, "edu": edu, "link": apply,
        "startDate": start, "endDate": end, "source": source,
        "date": collect_date, "note": note, "status": status
    })
with open('/workspace/docs/data.json', 'w', encoding='utf-8') as f:
    json.dump(json_data, f, ensure_ascii=False, indent=2)

print(f'已生成 {len(cats_sorted)+1} 个分类文件 + 大公司表 + data.json')
print(f'总岗位: {len(mapped_rows)} | 大公司: {len(big_items)}')