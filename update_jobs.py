#!/usr/bin/env python3
"""秋招投递追踪表更新脚本 - 2026-07-19"""

import json
import openpyxl
import requests
from datetime import datetime, date, timedelta
from copy import copy

TODAY = date.today()  # 2026-07-19
TODAY_STR = TODAY.strftime('%Y-%m-%d')
FILE_PATH = '/workspace/秋招投递追踪表.xlsx'
FEISHU_TOKEN = "t-g1047katZH6RYR4NO6N7P42VZZN5ARHMXFJLSH3G"
FEISHU_BASE = "Q6yQbgLiNa1y0as8QEtcOR6YnFb"
FEISHU_TABLE = "tblO4VkZJBoMZRrV"

# ========== 步骤1: 从飞书获取所有记录 ==========
def fetch_all_feishu_records():
    """分页获取飞书多维表格所有记录"""
    all_records = []
    page_token = None
    headers = {"Authorization": f"Bearer {FEISHU_TOKEN}"}
    
    while True:
        url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{FEISHU_BASE}/tables/{FEISHU_TABLE}/records?page_size=500"
        if page_token:
            url += f"&page_token={page_token}"
        
        resp = requests.get(url, headers=headers, timeout=30)
        data = resp.json()
        
        if data.get('code') != 0:
            print(f"Feishu API error: {data}")
            break
        
        items = data.get('data', {}).get('items', [])
        all_records.extend(items)
        
        has_more = data.get('data', {}).get('has_more', False)
        page_token = data.get('data', {}).get('page_token', '')
        
        print(f"  Fetched {len(items)} records, total so far: {len(all_records)}")
        if not has_more:
            break
    
    return all_records

# 泛商科关键词
BUSINESS_KEYWORDS = [
    '战略', '战投', '商业分析', '经营分析', '策略运营', '运营',
    '风险', '合规', '金融策略', '投资', '行研', '研究',
    '咨询', '产品', '销售', '市场', '营销', '财务', '审计',
    '管理培训生', '管培生', '管培', '培训生', '综合',
    '商务', '供应链', '采购', '人力', 'HR', '人力资源',
    '客户', '品牌', '策划', '分析师', '投行', '量化',
    '资管', '财富', '理财', '信贷', '风控', '法务',
    '基金', '证券', '银行', '保险', '信托', 'PE', 'VC',
    '经济', '金融', '会计', '税务', '审计', '数据分析',
    '项目经理', '产品经理', '运营管理', '市场推广',
    '投资管理', '资产管理', '风险管理', '信用管理',
    '业务', '职能', '解决方案', '行业研究', '公共事务',
    '政府关系', 'IR', 'PR', 'GR', '战略合作',
]

def is_business_role(job_title):
    """判断是否为泛商科岗位"""
    title_lower = job_title.lower()
    for kw in BUSINESS_KEYWORDS:
        if kw.lower() in title_lower:
            return True
    return False

def is_pure_intern(recruit_type):
    """判断是否为纯实习（排除）"""
    if not recruit_type:
        return False
    rt = recruit_type.strip()
    # 实习但非秋招提前批/校招的排除
    if '实习' in rt and '秋招提前批' not in rt and '校招' not in rt and '秋招' not in rt:
        return True
    if rt in ['实习', '暑期实习', '日常实习']:
        return True
    return False

def filter_feishu_records(records):
    """筛选飞书记录：27届、非春招、非纯实习、泛商科、更新时间>=2026-06"""
    filtered = []
    for item in records:
        fields = item.get('fields', {})
        
        # 招聘届别
        batch = fields.get('招聘届别', '')
        if not batch or '2027届' not in str(batch):
            continue
        
        # 招聘类型 - 排除春招
        recruit_type = fields.get('招聘类型', '')
        if not recruit_type:
            continue
        if '春招' in str(recruit_type) and '秋招' not in str(recruit_type):
            continue
        
        # 排除纯实习
        if is_pure_intern(str(recruit_type)):
            continue
        
        # 更新时间
        update_time = fields.get('更新时间', '')
        if update_time:
            try:
                ut = datetime.strptime(str(update_time)[:10], '%Y-%m-%d').date()
                if ut < date(2026, 6, 1):
                    continue
            except:
                pass
        
        # 泛商科岗位
        job_title = fields.get('招聘岗位', '')
        if not is_business_role(str(job_title)):
            continue
        
        filtered.append(fields)
    
    print(f"  Filtered to {len(filtered)} business-role records from {len(records)} total")
    return filtered

# ========== 步骤2: 从网络搜索获取新岗位 ==========
def get_web_search_jobs():
    """今日网络搜索发现的新岗位"""
    jobs = []
    
    # 百度2027届校园招聘 (全岗位)
    jobs.append({
        '公司名称': '百度',
        '所属行业': '互联网/科技',
        '岗位名称': '2027届校园招聘-产品类/商业分析/综合类',
        '岗位类别': '商业分析',
        '工作地点': '北京/上海/深圳/广州/杭州/成都',
        '学历要求': '硕士及以上',
        '毕业时间要求': '2027届',
        '网申开始日期': '2026-07-09',
        '网申截止日期': '招满为止',
        '投递方式/链接': 'https://talent.baidu.com/jobs',
        '信息来源': '网络搜索',
        '备注': '类型:秋招提前批; 届别:2027届; 74个细分职位; 5大方向'
    })
    
    # 华为2027届应届生-财经类
    jobs.append({
        '公司名称': '华为',
        '所属行业': '互联网/科技',
        '岗位名称': '2027届应届生-财经专员/税务专员/风控合规',
        '岗位类别': '财务/审计',
        '工作地点': '深圳/北京/上海',
        '学历要求': '硕士及以上',
        '毕业时间要求': '2027届',
        '网申开始日期': '2026-07-01',
        '网申截止日期': '招满为止',
        '投递方式/链接': 'https://career.huawei.com/',
        '信息来源': '网络搜索',
        '备注': '类型:校招; 届别:2027届; 财经类岗位; 面向2026.1-2027.12毕业生'
    })
    
    # 科大讯飞2027届常规校招-产品/职能
    jobs.append({
        '公司名称': '科大讯飞',
        '所属行业': '互联网/科技',
        '岗位名称': '2027届常规校招-产品类/职能类',
        '岗位类别': '产品',
        '工作地点': '合肥/北京/上海/广州/深圳',
        '学历要求': '硕士及以上',
        '毕业时间要求': '2027届',
        '网申开始日期': '2026-07-01',
        '网申截止日期': '招满为止',
        '投递方式/链接': 'https://campus.iflytek.com/',
        '信息来源': '网络搜索',
        '备注': '类型:校招; 届别:2027届; 技术类60%+产品类+职能类'
    })
    
    # 滴滴2027届秋招储备实习生
    jobs.append({
        '公司名称': '滴滴',
        '所属行业': '互联网/科技',
        '岗位名称': '2027届秋招储备实习生-商业分析/金融模型/产品/运营',
        '岗位类别': '商业分析',
        '工作地点': '北京/上海/杭州/广州',
        '学历要求': '硕士及以上',
        '毕业时间要求': '2027届',
        '网申开始日期': '2026-04-02',
        '网申截止日期': '招满为止',
        '投递方式/链接': 'https://campus.didiglobal.com/',
        '信息来源': '网络搜索',
        '备注': '类型:秋招提前批; 届别:2027届; 面向2026.9-2027.8毕业; 实习转正'
    })
    
    # 中信证券-宁波校招管培生
    jobs.append({
        '公司名称': '中信证券',
        '所属行业': '券商',
        '岗位名称': '校招培训生-宁波',
        '岗位类别': '金融策略',
        '工作地点': '宁波',
        '学历要求': '本科及以上',
        '毕业时间要求': '2027届',
        '网申开始日期': '2026-06-11',
        '网申截止日期': '2027-04-01',
        '投递方式/链接': 'https://careers.citics.com/',
        '信息来源': '网络搜索',
        '备注': '类型:校招; 届别:2027届; 面向2026下-2027上毕业'
    })
    
    # 中信证券-北京分公司培训生
    jobs.append({
        '公司名称': '中信证券',
        '所属行业': '券商',
        '岗位名称': '培训生-北京分公司',
        '岗位类别': '金融策略',
        '工作地点': '北京',
        '学历要求': '本科及以上',
        '毕业时间要求': '2027届',
        '网申开始日期': '2026-07-15',
        '网申截止日期': '招满为止',
        '投递方式/链接': 'https://careers.citics.com/',
        '信息来源': '网络搜索',
        '备注': '类型:校招; 届别:2027届'
    })
    
    # 中信证券-上海校招管培生
    jobs.append({
        '公司名称': '中信证券',
        '所属行业': '券商',
        '岗位名称': '校招管培生-上海',
        '岗位类别': '金融策略',
        '工作地点': '上海',
        '学历要求': '本科及以上',
        '毕业时间要求': '2027届',
        '网申开始日期': '2026-07-07',
        '网申截止日期': '招满为止',
        '投递方式/链接': 'https://careers.citics.com/',
        '信息来源': '网络搜索',
        '备注': '类型:校招; 届别:2027届'
    })
    
    # 美团2027届秋招提前批
    jobs.append({
        '公司名称': '美团',
        '所属行业': '互联网/科技',
        '岗位名称': '2027届秋招提前批-商业分析/产品/运营',
        '岗位类别': '商业分析',
        '工作地点': '北京/上海/深圳/成都',
        '学历要求': '硕士及以上',
        '毕业时间要求': '2027届',
        '网申开始日期': '2026-07-09',
        '网申截止日期': '招满为止',
        '投递方式/链接': 'https://campus.meituan.com/',
        '信息来源': '网络搜索',
        '备注': '类型:秋招提前批; 届别:2027届; 7月9日开启; 新增AI产品经理等岗位'
    })
    
    # 腾讯2027届实习生招聘(可转正)
    jobs.append({
        '公司名称': '腾讯',
        '所属行业': '互联网/科技',
        '岗位名称': '2027届实习生招聘-商业分析/产品/运营/市场(可转正)',
        '岗位类别': '商业分析',
        '工作地点': '深圳/北京/上海/广州',
        '学历要求': '硕士及以上',
        '毕业时间要求': '2027届',
        '网申开始日期': '2026-07-01',
        '网申截止日期': '招满为止',
        '投递方式/链接': 'https://join.qq.com/',
        '信息来源': '网络搜索',
        '备注': '类型:校招; 届别:2027届; 面向2026.9-2027.12毕业; 优秀实习生可留用'
    })
    
    # 快手-快Star
    jobs.append({
        '公司名称': '快手',
        '所属行业': '互联网/科技',
        '岗位名称': '快Star-2027届校园招聘-产品/运营/商业分析',
        '岗位类别': '商业分析',
        '工作地点': '北京/上海/杭州/深圳',
        '学历要求': '硕士及以上',
        '毕业时间要求': '2027届',
        '网申开始日期': '2026-07-14',
        '网申截止日期': '招满为止',
        '投递方式/链接': 'https://campus.kuaishou.cn/',
        '信息来源': '网络搜索',
        '备注': '类型:秋招提前批; 届别:2027届'
    })
    
    # 米哈游2027届提前批
    jobs.append({
        '公司名称': '米哈游',
        '所属行业': '互联网/科技',
        '岗位名称': '2027届秋招提前批-产品/运营/市场/商务',
        '岗位类别': '产品',
        '工作地点': '上海',
        '学历要求': '硕士及以上',
        '毕业时间要求': '2027届',
        '网申开始日期': '2026-07-14',
        '网申截止日期': '招满为止',
        '投递方式/链接': 'https://campus.mihoyo.com/',
        '信息来源': '网络搜索',
        '备注': '类型:秋招提前批; 届别:2027届'
    })
    
    # 欧莱雅2026暑期实习生(可直通2027管培生)
    # 已经在表中，跳过
    
    # 高盛2027届金融投行
    jobs.append({
        '公司名称': '高盛',
        '所属行业': '券商',
        '岗位名称': '2027届暑期分析师-投行/投资/研究/量化',
        '岗位类别': '投资/行研',
        '工作地点': '北京/上海/香港',
        '学历要求': '硕士及以上',
        '毕业时间要求': '2027届',
        '网申开始日期': '2026-07-15',
        '网申截止日期': '2026-10-15',
        '投递方式/链接': 'https://www.goldmansachs.com/careers/',
        '信息来源': '网络搜索',
        '备注': '类型:校招; 届别:2027届; 可转正'
    })
    
    print(f"  Web search: {len(jobs)} new candidate jobs")
    return jobs

# ========== 步骤3: 读取现有Excel ==========
def read_existing_excel():
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb['秋招投递追踪']
    
    existing = []
    existing_keys = set()
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        company = str(row[2]).strip() if row[2] else ''
        position = str(row[4]).strip() if row[4] else ''
        key = f'{company}|{position}'
        existing_keys.add(key)
        existing.append(row)
    
    return wb, ws, existing, existing_keys

# ========== 步骤4: 处理飞书记录并转换为表格格式 ==========
def convert_feishu_to_job(fields):
    """将飞书记录转换为表格格式"""
    company = fields.get('公司名称', '')
    job_title = fields.get('招聘岗位', '')
    
    # 确定行业类别
    industry = fields.get('行业类别', '其他')
    # 简化行业分类
    industry_map = {
        '证券': '券商', '基金': '基金', '银行': '银行', '保险': '保险',
        '信托': '信托', '期货': '券商', '投资': 'PE/VC',
        '互联网': '互联网/科技', '人工智能': '互联网/科技',
        '电商': '互联网/科技', '通信': '互联网/科技',
        '半导体': '互联网/科技', '电子': '制造业',
        '汽车': '制造业', '制造': '制造业', '新能源': '制造业',
        '快消': '快消', '零售': '快消', '餐饮': '快消',
        '咨询': '咨询', '教育': '咨询',
        '多元化': '制造业', '综合': '制造业',
        '航天': '制造业', '航空': '制造业', '军工': '制造业',
        '建筑': '制造业', '设计': '制造业',
        '医药': '制造业', '医疗': '制造业',
        '地产': '制造业', '物业': '制造业',
        '物流': '制造业', '交通': '制造业',
        '能源': '制造业', '化工': '制造业',
        '环保': '制造业', '农业': '制造业',
        '检测': '制造业', '测量': '制造业',
        '酒店': '快消', '旅游': '快消',
        '游戏': '互联网/科技', '娱乐': '互联网/科技',
        '软件': '互联网/科技', '数据': '互联网/科技',
        '机器人': '制造业',
    }
    
    for k, v in industry_map.items():
        if k in industry:
            industry = v
            break
    
    # 确定岗位类别
    title_lower = job_title.lower()
    if any(kw in title_lower for kw in ['战略', '战投']):
        category = '战略/战投'
    elif any(kw in title_lower for kw in ['商业分析', '经营分析', '数据分析']):
        category = '商业分析'
    elif any(kw in title_lower for kw in ['策略运营', '运营']):
        category = '策略运营'
    elif any(kw in title_lower for kw in ['风险', '合规', '风控']):
        category = '风险管理/合规'
    elif any(kw in title_lower for kw in ['投资', '行研', '研究', '量化', '投行']):
        category = '投资/行研'
    elif any(kw in title_lower for kw in ['咨询']):
        category = '咨询'
    elif any(kw in title_lower for kw in ['产品经理', '产品']):
        category = '产品'
    elif any(kw in title_lower for kw in ['销售', '市场', '营销', '商务', '客户', '品牌']):
        category = '销售/市场'
    elif any(kw in title_lower for kw in ['财务', '审计', '税务', '会计', '资金']):
        category = '财务/审计'
    elif any(kw in title_lower for kw in ['管培', '培训生', '综合', '职能', '人力', 'hr', '人力资源']):
        category = '管培/人力'
    elif any(kw in title_lower for kw in ['金融', '财富', '理财', '资管', '信贷', '证券']):
        category = '金融策略'
    elif any(kw in title_lower for kw in ['供应链', '采购']):
        category = '销售/市场'
    else:
        category = '其他'
    
    # 截止日期
    deadline = fields.get('截止日期', '招满为止')
    if not deadline:
        deadline = '招满为止'
    
    # 招聘类型
    recruit_type = fields.get('招聘类型', '')
    if '秋招提前批' in str(recruit_type):
        remark_type = '秋招提前批'
    else:
        remark_type = '校招'
    
    # 备注
    notes = f"类型:{remark_type}; 届别:2027届"
    if fields.get('笔试信息'):
        notes += f"; 笔试:{fields.get('笔试信息', '')}"
    
    return {
        '公司名称': company,
        '所属行业': industry,
        '岗位名称': job_title,
        '岗位类别': category,
        '工作地点': fields.get('工作地点', ''),
        '学历要求': '硕士及以上',
        '毕业时间要求': '2027届',
        '网申开始日期': fields.get('更新时间', ''),
        '网申截止日期': deadline,
        '投递方式/链接': fields.get('岗位链接', {}).get('link', '') if isinstance(fields.get('岗位链接'), dict) else str(fields.get('岗位链接', '')),
        '信息来源': '飞书秋招合集',
        '备注': notes
    }

# ========== 主处理流程 ==========
def main():
    print("=" * 60)
    print(f"秋招投递追踪表更新 - {TODAY_STR}")
    print("=" * 60)
    
    # 读取现有表格
    print("\n[1] 读取现有表格...")
    wb, ws, existing_data, existing_keys = read_existing_excel()
    max_seq = max([row[0] for row in existing_data if row[0] and isinstance(row[0], (int, float))], default=0)
    print(f"  现有 {len(existing_data)} 条记录, 最大序号 {max_seq}")
    
    # 从飞书获取数据
    print("\n[2] 从飞书获取数据...")
    feishu_records = fetch_all_feishu_records()
    feishu_filtered = filter_feishu_records(feishu_records)
    
    # 转换飞书数据
    feishu_jobs = [convert_feishu_to_job(f) for f in feishu_filtered]
    print(f"  转换后飞书岗位: {len(feishu_jobs)}")
    
    # 从网络搜索获取数据
    print("\n[3] 从网络搜索获取新岗位...")
    web_jobs = get_web_search_jobs()
    
    # 合并所有新岗位
    all_new_jobs = feishu_jobs + web_jobs
    print(f"\n  合并后候选岗位: {len(all_new_jobs)}")
    
    # 去重
    print("\n[4] 去重与筛选...")
    unique_jobs = []
    new_keys = set()
    today_deadline = 0
    near_deadline = []
    
    for job in all_new_jobs:
        company = job['公司名称'].strip()
        position = job['岗位名称'].strip()
        key = f'{company}|{position}'
        
        # 跳过已存在的
        if key in existing_keys or key in new_keys:
            continue
        
        # 检查截止日期
        deadline = job['网申截止日期']
        if deadline and deadline != '招满为止':
            try:
                dl = datetime.strptime(str(deadline)[:10], '%Y-%m-%d').date()
                if dl < TODAY:
                    # 已过期，跳过
                    continue
                elif dl == TODAY:
                    job['备注'] = (job.get('备注', '') + '; 今日截止！').strip('; ')
                    today_deadline += 1
                elif dl <= TODAY + timedelta(days=7):
                    job['备注'] = (job.get('备注', '') + '; 即将截止').strip('; ')
                    near_deadline.append(f"{company}-{position}({deadline})")
            except:
                pass
        
        new_keys.add(key)
        unique_jobs.append(job)
    
    print(f"  去重后新岗位: {len(unique_jobs)}")
    print(f"  今日截止: {today_deadline}")
    if near_deadline:
        print(f"  即将截止: {near_deadline}")
    
    # 统计
    beijing_count = sum(1 for j in unique_jobs if '北京' in str(j['工作地点']))
    shanghai_count = sum(1 for j in unique_jobs if '上海' in str(j['工作地点']))
    companies = set(j['公司名称'] for j in unique_jobs)
    industries = set(j['所属行业'] for j in unique_jobs)
    
    print(f"\n  新增岗位统计:")
    print(f"    - 共 {len(unique_jobs)} 个新岗位")
    print(f"    - 来自 {len(companies)} 家公司")
    print(f"    - 北京 {beijing_count} 个, 上海 {shanghai_count} 个")
    print(f"    - 主要行业: {', '.join(industries)}")
    
    if len(unique_jobs) == 0:
        print("\n  没有新岗位需要追加，跳过写入。")
        # 仍然更新状态
        print("\n[7] 更新岗位状态...")
        update_statuses(ws, existing_data)
        wb.save(FILE_PATH)
        print("  状态更新完成。")
        return
    
    # 追加新记录
    print(f"\n[5] 追加 {len(unique_jobs)} 条新记录...")
    seq = int(max_seq)
    
    for job in unique_jobs:
        seq += 1
        row_data = [
            seq,
            TODAY_STR,
            job['公司名称'],
            job['所属行业'],
            job['岗位名称'],
            job['岗位类别'],
            job['工作地点'],
            job['学历要求'],
            job['毕业时间要求'],
            job['网申开始日期'],
            job['网申截止日期'],
            job['投递方式/链接'],
            job['信息来源'],
            '否',  # 是否已投递
            None,  # 投递日期
            '待投递',  # 当前状态
            job.get('备注', '')
        ]
        ws.append(row_data)
    
    total_count = len(existing_data) + len(unique_jobs)
    open_count = sum(1 for r in existing_data if r[15] == '待投递') + len(unique_jobs)
    print(f"  追加完成。累计 {total_count} 条记录。")
    
    # 更新每日日志
    print("\n[6] 更新每日日志...")
    ws2 = wb['每日更新日志']
    feishu_count = len([j for j in unique_jobs if j['信息来源'] == '飞书秋招合集'])
    web_count = len([j for j in unique_jobs if j['信息来源'] == '网络搜索'])
    sources = []
    if feishu_count > 0:
        sources.append(f"飞书秋招合集({feishu_count})")
    if web_count > 0:
        sources.append(f"网络搜索({web_count})")
    
    ws2.append([
        TODAY_STR,
        len(unique_jobs),
        today_deadline,
        total_count,
        open_count,
        ', '.join(sources),
        '自动更新'
    ])
    print(f"  日志已更新。")
    
    # 更新状态
    print("\n[7] 更新岗位状态...")
    update_statuses(ws, existing_data)
    
    # 保存
    wb.save(FILE_PATH)
    print(f"  文件已保存: {FILE_PATH}")
    
    print("\n" + "=" * 60)
    print("更新完成！")
    print("=" * 60)
    
    # 输出摘要
    print(f"\n今天新增{len(unique_jobs)}个岗位，来自{len(companies)}家公司，主要方向是{', '.join(industries)}。"
          f"其中北京{beijing_count}个、上海{shanghai_count}个。当前累计共{total_count}个岗位。")
    if near_deadline:
        print(f"\n临近截止提醒:")
        for nd in near_deadline:
            print(f"  - {nd}")
    print(f"\nHTML仪表盘: https://karis5208.github.io/seek-job/")
    print(f"GitHub仓库: https://github.com/Karis5208/seek-job")

def update_statuses(ws, existing_data):
    """更新过期岗位状态"""
    updated = 0
    for row_idx, row_data in enumerate(existing_data, start=2):
        deadline = row_data[10]  # 网申截止日期
        status = row_data[15]  # 当前状态
        
        if status != '待投递':
            continue
        
        if deadline and deadline != '招满为止':
            try:
                dl = datetime.strptime(str(deadline)[:10], '%Y-%m-%d').date()
                if dl < TODAY:
                    ws.cell(row=row_idx, column=16).value = '已截止'
                    updated += 1
            except:
                pass
    
    print(f"  已更新 {updated} 条过期记录状态为'已截止'。")

if __name__ == '__main__':
    main()