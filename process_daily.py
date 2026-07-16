#!/usr/bin/env python3
"""秋招信息每日追踪 - 完整处理脚本"""
import json, os, re, subprocess
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill

TODAY = datetime.now()
TODAY_STR = TODAY.strftime("%Y-%m-%d")
EXCEL_PATH = "/workspace/秋招投递追踪表.xlsx"

# ============================================================
# 泛商科岗位关键词
# ============================================================
BUSINESS_KEYWORDS = [
    "战略", "投资", "商业分析", "经营分析", "运营", "策略",
    "风险", "合规", "金融", "行研", "研究", "咨询", "产品",
    "销售", "市场", "财务", "审计", "管培", "管理培训", "风控",
    "量化", "资管", "财富管理", "投行", "投研", "交易", "商务",
    "营销", "品牌", "公关", "客户", "数据", "企划", "采购",
    "培训", "招聘", "HR", "人力", "行政", "助理", "秘书",
    "法务", "内控", "精算", "信评", "评级", "固收", "权益",
    "宏观", "策略研究", "研究员", "分析师", "财富", "理财",
    "资产", "债券", "外汇", "衍生品", "经济", "会计",
    "税务", "审计", "总助", "CEO", "总裁", "供应链",
    "物流", "贸易", "零售", "传媒", "内容", "编辑",
    "直播", "视频", "策划", "广告", "增长", "BD",
    "渠道", "贸易", "供应链", "SCM", "MT", "培训生",
    "储备干部", "综合管理", "新媒体", "电商", "用户研究",
    "行业研究", "市场研究", "BI", "商业智能", "产品经理",
    "PM", "consulting", "advisory", "解决方案",
]

FINANCE_INDUSTRIES = ["金融", "证券", "银行", "保险", "投资", "基金", "信托", "期货", "租赁", "担保", "支付", "资产管理", "财富管理", "审计", "财税", "会计师事务所", "融资", "国企", "央企"]

# ============================================================
# 岗位分类
# ============================================================
def classify_job(company, job_title, industry):
    text = str(company or "") + " " + str(job_title or "")
    rules = [
        (["金融", "银行", "证券", "保险", "投资", "基金", "信托", "期货", "融资", "信贷", "风控", "风险", "合规", "法务", "精算", "信评", "评级", "投行", "量化", "交易", "财富", "理财", "资管", "资产", "债券", "外汇", "利率", "衍生品", "股票", "研究员", "分析师", "行研", "固收", "权益", "宏观", "策略研究"], "金融"),
        (["战略", "策略", "企划", "规划"], "战略"),
        (["商业分析", "商分", "经营分析", "业务分析", "数据分析", "数据科学", "BI", "商业智能", "用户研究", "行业研究", "市场研究", "市场分析"], "商分"),
        (["运营", "内容运营", "用户运营", "活动运营", "新媒体运营", "社群运营", "电商运营", "平台运营", "产品运营", "增长运营"], "运营"),
        (["管培", "管培生", "培训生", "储备干部", "管理培训", "MT", "综合管理"], "管培"),
        (["市场", "营销", "品牌", "推广", "公关", "商务", "渠道", "销售", "客户", "广告", "策划", "增长", "BD", "GTM", "新媒体"], "市场"),
        (["产品", "产品经理", "产品设计", "产品策划", "产品助理", "PM"], "产品"),
        (["咨询", "顾问", "consulting", "advisory", "解决方案"], "咨询"),
        (["人力", "HR", "人事", "招聘", "薪酬", "培训", "组织发展", "OD", "员工关系", "行政", "秘书", "总助", "助理", "财务", "会计", "税务", "审计", "法务", "合规", "内控", "采购"], "职能"),
        (["供应链", "采购", "物流", "仓储", "SCM", "供应链管理", "贸易"], "供应链"),
        (["传媒", "媒体", "内容", "编辑", "记者", "视频", "直播", "文案", "设计", "UI", "UX", "视觉", "创意", "编导", "出版"], "传媒"),
        (["零售", "门店", "店铺", "陈列", "导购", "店长", "卖场"], "零售"),
    ]
    for keywords, category in rules:
        for kw in keywords:
            if kw in text:
                return category
    return "其他"

# ============================================================
# 读取现有Excel
# ============================================================
def read_existing():
    if not os.path.exists(EXCEL_PATH):
        return []
    wb = load_workbook(EXCEL_PATH)
    if "秋招投递追踪" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["秋招投递追踪"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(list(row))
    wb.close()
    print(f"  现有Excel记录: {len(rows)} 条")
    return rows

# ============================================================
# 读取Feishu数据并筛选
# ============================================================
def read_feishu():
    with open('/workspace/feishu_raw.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    items = data['items']
    
    filtered = []
    for item in items:
        fields = item['fields']
        batch = str(fields.get('招聘届别', ''))
        job_type = str(fields.get('招聘类型', ''))
        company = str(fields.get('公司名称', '')).strip()
        positions = str(fields.get('招聘岗位', '')).strip()
        industry = str(fields.get('行业类别', '')).strip()
        location = str(fields.get('工作地点', '')).strip()
        deadline = str(fields.get('截止日期', '')).strip()
        job_link = fields.get('岗位链接', {})
        announce_link = fields.get('招聘公告', {})
        exam_info = str(fields.get('笔试信息', '')).strip()
        company_size = str(fields.get('公司规模', '')).strip()
        start_date = str(fields.get('更新时间', '')).strip()
        
        # 筛选27届
        if '27' not in batch:
            continue
        # 排除春招
        if '春招' in job_type:
            continue
        
        # 泛商科筛选
        pos_str = str(positions)
        matched = any(kw in pos_str for kw in BUSINESS_KEYWORDS)
        if not matched:
            ind_str = str(industry)
            if any(kw in ind_str for kw in FINANCE_INDUSTRIES):
                matched = True
            elif any(kw in ind_str for kw in ['互联网', '咨询', '快消', '消费', '人工智能']):
                matched = any(kw in pos_str for kw in BUSINESS_KEYWORDS)
        if not matched:
            continue
        
        # 获取投递链接
        apply_link = ""
        if isinstance(job_link, dict) and job_link.get("link"):
            apply_link = job_link["link"]
        elif isinstance(announce_link, dict) and announce_link.get("link"):
            apply_link = announce_link["link"]
        
        # 分类
        category = classify_job(company, positions, industry)
        
        # 毕业时间要求
        grad_req = "2027届"
        if "2026" in batch and "2027" in batch:
            grad_req = "2026届/2027届"
        elif "2025" in batch:
            grad_req = "2025-2027届"
        
        # 备注
        notes = []
        if job_type:
            notes.append(f"类型:{job_type}")
        if batch:
            notes.append(f"届别:{batch}")
        if exam_info:
            notes.append(f"笔试:{exam_info}")
        if company_size:
            notes.append(f"规模:{company_size}")
        note_str = "; ".join(notes)
        
        filtered.append({
            'company': company,
            'industry': industry,
            'position': positions,
            'category': category,
            'location': location,
            'edu': '硕士及以上',
            'grad_req': grad_req,
            'start_date': start_date,
            'deadline': deadline,
            'apply_link': apply_link,
            'source': '飞书秋招合集',
            'note': note_str,
        })
    
    print(f"  飞书原始: {len(items)} 条, 筛选后: {len(filtered)} 条")
    return filtered

# ============================================================
# 去重
# ============================================================
def deduplicate(existing, new_records):
    existing_keys = set()
    for row in existing:
        company = str(row[2] or "").strip()
        position = str(row[4] or "").strip()[:30]
        existing_keys.add(company + "|||" + position)
    
    truly_new = []
    for rec in new_records:
        key = rec['company'] + "|||" + rec['position'][:30]
        if key not in existing_keys:
            truly_new.append(rec)
    
    print(f"  去重: 新记录 {len(new_records)} -> 去重后 {len(truly_new)} 条")
    return truly_new

# ============================================================
# 检查截止日期
# ============================================================
def check_deadline(deadline_str):
    """返回 (is_valid, deadline_date, note)"""
    if not deadline_str or deadline_str in ["招满为止", "不限", ""]:
        return True, None, ""
    
    dl = deadline_str.strip().replace(".", "-").replace("/", "-")
    try:
        dl_date = datetime.strptime(dl[:10], "%Y-%m-%d")
        days_left = (dl_date - TODAY).days
        
        if days_left < 0:
            return False, dl_date, "已截止"
        elif days_left == 0:
            return True, dl_date, "今日截止！"
        elif days_left <= 7:
            return True, dl_date, f"即将截止({days_left}天后)"
        else:
            return True, dl_date, ""
    except:
        return True, None, ""

# ============================================================
# 更新Excel
# ============================================================
def update_excel(existing_rows, new_records):
    wb = load_workbook(EXCEL_PATH) if os.path.exists(EXCEL_PATH) else Workbook()
    
    # 确保sheet存在
    sheet_name = "秋招投递追踪"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
        ws.title = sheet_name
    
    # 清空并重建（保留现有数据+新增）
    # 先读取所有现有行
    all_rows = []
    for row in existing_rows:
        all_rows.append(list(row))
    
    # 新记录转换为Excel行
    new_excel_rows = []
    start_seq = len(all_rows) + 1
    for i, rec in enumerate(new_records):
        is_valid, dl_date, dl_note = check_deadline(rec['deadline'])
        if not is_valid:
            continue  # 已截止的跳过
        
        # 合并备注
        final_note = rec['note']
        if dl_note:
            final_note = final_note + "; " + dl_note if final_note else dl_note
        
        # 学历要求
        edu = rec['edu']
        if '硕士' not in edu and '博士' not in edu:
            edu = '硕士及以上'
        
        new_excel_rows.append([
            start_seq + i,
            TODAY_STR,
            rec['company'],
            rec['industry'],
            rec['position'],
            rec['category'],
            rec['location'],
            edu,
            rec['grad_req'],
            rec['start_date'],
            rec['deadline'],
            rec['apply_link'],
            rec['source'],
            '否',
            '',
            '待投递',
            final_note,
        ])
    
    # 合并
    all_rows.extend(new_excel_rows)
    
    # 写入表头
    headers = ["序号", "记录日期", "公司名称", "所属行业", "岗位名称", "岗位类别", "工作地点", "学历要求", "毕业时间要求", "网申开始日期", "网申截止日期", "投递方式/链接", "信息来源", "是否已投递", "投递日期", "当前状态", "备注"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 写入数据
    for row_idx, row_data in enumerate(all_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value or "")
    
    # 调整列宽
    widths = [6, 12, 22, 14, 38, 10, 20, 10, 14, 12, 12, 40, 15, 10, 10, 10, 30]
    for i, w in enumerate(widths, 1):
        col_letter = chr(64 + i) if i <= 26 else chr(64 + (i - 1) // 26) + chr(65 + (i - 1) % 26)
        ws.column_dimensions[col_letter].width = w
    
    ws.freeze_panes = "A2"
    
    # 更新每日日志sheet
    update_log_sheet(wb, len(new_excel_rows), len(all_rows))
    
    wb.save(EXCEL_PATH)
    wb.close()
    print(f"  Excel已保存: 总 {len(all_rows)} 条, 新增 {len(new_excel_rows)} 条")
    
    return len(new_excel_rows), len(all_rows), all_rows

def update_log_sheet(wb, new_count, total_count):
    sheet_name = "每日更新日志"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        headers = ["日期", "新增岗位数", "当日截止岗位数", "累计岗位总数", "开放中数量", "主要来源", "备注"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
    
    # 找到最后一行
    last_row = ws.max_row + 1
    ws.cell(row=last_row, column=1, value=TODAY_STR)
    ws.cell(row=last_row, column=2, value=new_count)
    ws.cell(row=last_row, column=3, value=0)
    ws.cell(row=last_row, column=4, value=total_count)
    ws.cell(row=last_row, column=5, value=total_count)  # 开放中数量后续更新
    ws.cell(row=last_row, column=6, value="飞书秋招合集+WebSearch")
    ws.cell(row=last_row, column=7, value="自动更新")

# ============================================================
# 更新过期状态
# ============================================================
def update_expired():
    wb = load_workbook(EXCEL_PATH)
    ws = wb["秋招投递追踪"]
    updated = 0
    for row in ws.iter_rows(min_row=2):
        status_cell = row[15]  # P列: 当前状态
        deadline_cell = row[10]  # K列: 网申截止日期
        if status_cell.value == "已截止":
            continue
        deadline = str(deadline_cell.value or "").strip()
        if not deadline or deadline in ["招满为止", "不限", "None", ""]:
            continue
        dl = deadline.replace(".", "-").replace("/", "-")
        try:
            dl_date = datetime.strptime(dl[:10], "%Y-%m-%d")
            if dl_date < TODAY:
                status_cell.value = "已截止"
                updated += 1
        except:
            pass
    
    if updated:
        wb.save(EXCEL_PATH)
        print(f"  标记已截止: {updated} 条")
    wb.close()
    return updated

# ============================================================
# 统计信息
# ============================================================
def get_stats(all_rows):
    total = len(all_rows)
    open_count = sum(1 for r in all_rows if r[15] != "已截止")
    today_deadline = []
    soon_deadline = []
    
    for row in all_rows:
        deadline = str(row[10] or "").strip()
        if not deadline or deadline in ["招满为止", "不限", "None", ""]:
            continue
        dl = deadline.replace(".", "-").replace("/", "-")
        try:
            dl_date = datetime.strptime(dl[:10], "%Y-%m-%d")
            days = (dl_date - TODAY).days
            if days == 0:
                today_deadline.append(row)
            elif 0 < days <= 7:
                soon_deadline.append((row, days))
        except:
            pass
    
    return total, open_count, today_deadline, soon_deadline

# ============================================================
# 生成Markdown
# ============================================================
def generate_markdown():
    wb = load_workbook(EXCEL_PATH)
    ws = wb["秋招投递追踪"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    data_rows = rows[1:]  # 跳过表头
    # 只展示未截止的
    open_rows = [r for r in data_rows if r[15] != "已截止"]
    
    # 按类别分组
    by_category = {}
    for row in open_rows:
        cat = row[5] if row[5] else "其他"
        if cat not in by_category:
            by_category[cat] = []
        by_category[cat].append(row)
    
    cat_order = ["金融", "战略", "商分", "运营", "管培", "市场", "产品", "咨询", "职能", "供应链", "传媒", "零售"]
    cats_sorted = sorted(by_category.items(), key=lambda x: (cat_order.index(x[0]) if x[0] in cat_order else 99, -len(x[1])))
    
    # 大公司名单
    big_companies = {
        "工商银行", "建设银行", "农业银行", "中国银行", "交通银行", "邮储银行",
        "招商银行", "兴业银行", "浦发银行", "中信银行", "民生银行", "平安银行",
        "宁波银行", "南京银行", "杭州银行", "上海银行", "北京银行",
        "中信证券", "中金公司", "华泰证券", "国泰君安", "海通证券", "申万宏源",
        "招商证券", "广发证券", "中国银河", "中信建投", "国信证券", "东方证券",
        "易方达", "华夏基金", "广发基金", "南方基金", "富国基金", "嘉实基金",
        "博时基金", "天弘基金", "汇添富", "泰康资产",
        "中国人寿", "中国平安", "中国人保", "中国太保",
        "腾讯", "阿里巴巴", "字节跳动", "百度", "美团", "京东", "拼多多",
        "网易", "小米", "快手", "滴滴", "小红书", "B站", "携程",
        "大疆", "华为", "联想", "OPPO", "vivo", "蚂蚁集团",
        "宝洁", "联合利华", "欧莱雅", "雀巢", "百事", "可口可乐", "玛氏",
        "亿滋", "蒙牛", "伊利", "安踏", "李宁", "耐克",
        "普华永道", "德勤", "安永", "毕马威", "埃森哲",
        "蔚来", "理想", "小鹏", "比亚迪", "宁德时代", "特斯拉",
        "华润", "中粮", "中化", "中国移动", "中国电信", "中国联通",
        "国家电网", "南方电网", "万科", "保利", "龙湖", "绿城",
        "招商蛇口", "中海", "华润置地", "美的", "格力", "海尔",
        "TCL", "海信", "顺丰", "科大讯飞", "海康威视", "大华",
        "米哈游", "莉莉丝", "叠纸", "鹰角", "SHEIN", "希音",
        "中金", "CICC", "中国烟草", "中国建筑", "中国中铁", "中国铁建",
        "中国交建", "三峡集团", "招商局", "中国船舶", "中国航天",
        "中国商飞", "中国航发", "中国电建", "中国能建", "中核集团",
        "中国广核", "中国兵器", "中国电子", "中国电科", "中国信科",
        "中国邮政", "中国国新", "中国诚通", "中国物流", "中国医药",
        "国药集团", "上海医药", "广州医药", "九州通", "恒瑞医药",
        "药明康德", "百济神州", "迈瑞医疗", "联影", "华大",
        "中国神华", "中煤", "中国石油", "中国石化", "中国海油",
        "紫金矿业", "中国铝业", "中国五矿", "中国黄金", "中国稀土",
        "上汽集团", "一汽", "东风", "长安", "吉利", "长城", "奇瑞",
        "广汽", "北汽", "华晨", "江铃", "中国重汽", "陕汽",
        "中国旅游", "中国中免", "中免", "华侨城", "中国国航",
        "中国南方航空", "中国东方航空", "海航", "春秋", "吉祥",
        "上海机场", "白云机场", "深圳机场", "首都机场",
        "港交所", "上交所", "深交所", "北交所", "中国结算",
        "上海清算所", "中国金融期货", "大连商品", "郑州商品",
        "银河证券", "中信建投", "中泰证券", "光大证券", "兴业证券",
        "东方财富", "东方证券", "国金证券", "浙商证券", "天风证券",
        "方正证券", "东吴证券", "国海证券", "西部证券", "国联证券",
        "华安证券", "长城证券", "国元证券", "西南证券", "东北证券",
        "财通证券", "华西证券", "山西证券", "第一创业", "国盛证券",
        "开源证券", "万联证券", "华金证券", "华兴证券", "东亚前海",
        "汇丰前海", "野村东方", "大和证券", "瑞银证券", "高盛高华",
        "摩根士丹利证券", "摩根大通证券", "星展证券", "渣打证券",
        "招商基金", "中欧基金", "兴证全球", "景顺长城", "鹏华基金",
        "工银瑞信", "银华基金", "中庚基金", "拾贝投资", "启林投资",
        "东方红", "睿远", "泉果", "高毅", "景林", "淡水泉",
        "源乐晟", "重阳", "星石", "千合", "仁桥", "慎知",
        "聚鸣", "磐耀", "丹羿", "同犇", "望正", "线性资本",
        "九合创投", "中联基金", "平方和投资", "信达证券", "华鑫证券",
        "财达证券", "华林证券", "中航证券", "红塔证券", "国都证券",
        "联储证券", "首创证券", "民生证券", "华创证券", "南京证券",
        "泰康保险", "阳光保险", "新华保险", "中银证券", "太平洋证券",
        "华润银行", "北京农商银行", "上海农商银行", "深圳农商银行",
        "广州农商银行", "成都银行", "重庆银行", "苏州银行",
        "长沙银行", "郑州银行", "青岛银行", "厦门银行", "齐鲁银行",
        "兰州银行", "西安银行", "贵阳银行", "东莞银行", "圆通",
        "货拉拉", "京东方", "超聚变", "微步在线", "多益网络",
        "商汤科技", "虹软科技", "新和成", "绿城中国", "恩智浦",
        "理想汽车", "蔚来", "小鹏汽车", "阶跃星辰", "智元机器人",
        "货拉拉", "58同城", "网易游戏雷火", "点点互动",
        "腾讯-青云计划", "字节跳动-AI产品经理", "百度-管理培训生",
        "京东-TET管理培训生", "拼多多集团-PDD", "优衣库",
        "华金证券", "仲信国际融资租赁", "中国电信",
        "华润银行", "南芯科技", "芯动科技", "矽力杰",
        "BrainCo强脑科技", "乐鑫科技", "思瑞浦", "江森自控",
        "燧原科技", "中国石油化工", "中国航天科工", "中国航天科技",
        "中航", "中船", "中兵", "兵工", "兵装",
        "中国电子科技", "中国电子信息", "国机集团", "通用技术",
        "中国建材", "中国有色", "矿冶科技", "中国中冶", "中冶",
        "中国煤炭科工", "中国煤炭地质", "中国钢研", "有研",
        "中国化学", "中国化学工程", "中化集团", "中国化工",
        "中国节能", "中国环保", "中国林业", "中国农垦",
        "中国农发", "中国供销", "中国再生资源", "中国盐业",
        "中国纺织", "中国轻工", "中国食品", "中国工艺",
        "中国生物", "中国中药", "中国医药集团", "中国出版",
        "中国电影", "中国广播电视", "中国有线", "中国广电",
        "中国卫通", "中国信通", "中国铁塔", "中国通号",
        "中国信保", "中国贸促", "中国期货", "中国证券",
        "中国基金", "中国保险", "中国信托", "中国融资担保",
        "中国小额贷款", "中国融资租赁", "中国商业保理",
        "中国上市公司", "中国并购", "中国总会计师",
        "中国注册会计师", "中国资产评估", "中国注册税务师",
        "中国保险学会", "中国金融学会", "中国城市金融",
        "中国农村金融", "中国国际金融", "中国金融40人",
        "上海金融", "深圳金融", "北京金融", "广州金融",
        "中国电力", "华能", "华电", "大唐", "国电", "中电投",
        "中国绿发", "鲁能", "华润电力", "国投电力",
        "上海证券交易所", "深圳证券交易所", "中国金融期货交易所",
        "北京证券交易所", "全国股转", "股转公司",
        "中国证券登记结算", "中国证券金融",
        "社保基金", "中投公司", "中国投资", "丝路基金",
        "国开金融", "建银投资", "信达资产", "华融资管",
        "长城资产", "东方资产", "中国东方",
        "外汇管理局", "外汇交易中心", "中央结算",
        "银联", "网联", "跨境清算", "人民币跨境支付",
        "微众银行", "网商银行", "新网银行", "百信银行",
        "苏宁银行", "亿联银行", "中关村银行", "众邦银行",
        "BlueFocus", "蓝色光标", "奥美", "阳狮", "WPP",
        "电通", "IPG", "宏盟", "HAVAS", "汉威士",
        "博报堂", "ADK", "McCann", "麦肯", "BBDO",
        "DDB", "TBWA", "李奥贝纳", "Leo Burnett",
        "来赞达", "Lazada", "Shopee", "TikTok", "抖音",
        "Shein", "希音", "Temu", "Alibaba", "阿里",
        "Tencent", "ByteDance", "Meituan", "Pinduoduo",
        "Xiaomi", "Huawei", "DJI", "DJI大疆", "Bilibili",
        "Kuaishou", "Zhihu", "Xiaohongshu", "Didi",
        "Manbang", "Huolala", "Hello", "哈啰",
        "Ctrip", "Trip.com", "Qunar", "Tongcheng",
        "Keep", "Meitu", "WPS", "金山办公", "Yongyou",
        "Kingdee", "Inspur", "ZTE", "Hikvision",
        "Dahua", "Uniview", "宇视", "Tiandy", "天地伟业",
        "IFLYTEK", "科大讯飞", "SenseTime", "商汤",
        "Megvii", "旷视", "YITU", "依图", "CloudWalk",
        "4Paradigm", "第四范式", "MiningLamp", "明略",
        "Horizon", "地平线", "BlackSesame", "黑芝麻",
        "Cambricon", "寒武纪", "Enflame", "燧原",
        "Biren", "壁仞", "MooreThreads", "摩尔线程",
        "Iluvatar", "天数智芯", "Vastai", "瀚博",
        "MetaX", "沐曦", "Denglin", "登临",
        "Hygon", "海光", "Loongson", "龙芯", "Phytium", "飞腾",
        "Zhaoxin", "兆芯", "Kunpeng", "鲲鹏", "Ascend", "昇腾",
        "Kirin", "麒麟", "UnionTech", "统信", "UOS",
        "Deepin", "深度", "Euler", "欧拉", "OpenEuler",
        "Dameng", "达梦", "Kingbase", "人大金仓", "GBase",
        "南大通用", "Shenzhou", "神舟通用", "UXsino", "优炫",
        "Hundsun", "恒生电子", "Apex", "顶点", "Kingdom", "金证",
        "Yucheng", "赢时胜", "Sunline", "长亮", "GlobalInfo",
        "科蓝", "Yusys", "宇信", "Tansun", "天阳", "SinoSoft",
        "中科软", "Hoperun", "润和", "BeyondSoft", "博彦",
        "Chinasoft", "中软国际", "VanceInfo", "文思海辉",
        "iSoftStone", "软通动力", "Neusoft", "东软",
        "Hand", "汉得", "Sie", "赛意", "Digiwin", "鼎捷",
        "Mysoft", "明源", "Glodon", "广联达", "Shiji", "石基",
    }
    
    def is_big(company):
        company = str(company)
        for bc in big_companies:
            if bc in company:
                return True
        return False
    
    # 主索引
    lines = [
        '# 2026 秋招投递追踪',
        '',
        f'> 每天早上自动更新 | 最后更新：{TODAY_STR} | 共 {len(open_rows)} 条开放岗位',
        '',
        '📱 手机和网页端均可直接查看。',
        '',
        '## 大公司精选',
        '',
        f'[大公司投递表](./docs/大公司投递表.md) — 跨行业大厂/国企/央企精选',
        '',
        '## 岗位分类',
        '',
        '| 分类 | 数量 |',
        '|------|------|',
    ]
    
    for cat, items in cats_sorted:
        safe_name = cat.replace("/", "-")
        lines.append(f'| [{cat}](./docs/{safe_name}.md) | {len(items)} 条 |')
    
    lines += [
        '',
        '---',
        '',
        '同时提供 [Excel版本](./秋招投递追踪表.xlsx)，下载后可筛选排序、标记已投递。',
        '',
        '*每日自动更新 · 祝秋招顺利！*',
    ]
    
    with open('/workspace/autumn_recruitment.md', 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    
    # 各分类文件
    os.makedirs('/workspace/docs', exist_ok=True)
    for cat, items in cats_sorted:
        safe_name = cat.replace("/", "-")
        items_sorted = sorted(items, key=lambda r: (not is_big(r[2]), r[15] != "待投递"))
        
        clines = [f'# {cat}']
        clines.append('')
        clines.append(f'> 共 {len(items)} 条 | 更新于 {TODAY_STR}')
        clines.append('')
        clines.append('| 公司 | 岗位 | 地点 | 截止 | 投递 | 备注 | 状态 |')
        clines.append('|------|------|------|------|------|------|------|')
        
        for row in items_sorted:
            company = str(row[2] or '—')[:20]
            job = str(row[4] or '—')[:35]
            location = str(row[6] or '—')[:12]
            deadline = str(row[10] or '—')[:10]
            apply_link = row[11] or ''
            note = str(row[16] or '—')[:20]
            status_val = str(row[15] or '待投递')
            status_e = '🟢' if status_val == '待投递' else '🔴'
            apply_cell = f'[投递]({apply_link})' if apply_link and apply_link.startswith('http') else '—'
            clines.append(f'| {company} | {job} | {location} | {deadline} | {apply_cell} | {note} | {status_e} {status_val} |')
        
        clines.append('')
        clines.append(f'[← 返回首页](../autumn_recruitment.md)')
        
        with open(f'/workspace/docs/{safe_name}.md', 'w', encoding='utf-8') as f:
            f.write('\n'.join(clines))
    
    # 大公司专属表
    big_items = [r for r in open_rows if is_big(r[2])]
    big_by_cat = {}
    for row in big_items:
        cat = row[5] if row[5] else "其他"
        if cat not in big_by_cat:
            big_by_cat[cat] = []
        big_by_cat[cat].append(row)
    
    big_cats_sorted = sorted(big_by_cat.items(), key=lambda x: (cat_order.index(x[0]) if x[0] in cat_order else 99, -len(x[1])))
    
    blines = [
        '# 大公司投递表',
        '',
        f'> 精选大厂/国企/央企/知名企业 | 更新于 {TODAY_STR} | 共 {len(big_items)} 条',
        '',
        '[← 返回全部岗位](../autumn_recruitment.md) | [Excel版本](../秋招投递追踪表.xlsx)',
        '',
    ]
    
    for cat, items in big_cats_sorted:
        items_sorted = sorted(items, key=lambda r: (r[15] != "待投递"))
        blines.append(f'## {cat} ({len(items)}条)')
        blines.append('')
        blines.append('| 公司 | 岗位 | 地点 | 截止 | 投递 | 备注 | 状态 |')
        blines.append('|------|------|------|------|------|------|------|')
        
        for row in items_sorted:
            company = str(row[2] or '—')[:20]
            job = str(row[4] or '—')[:35]
            location = str(row[6] or '—')[:12]
            deadline = str(row[10] or '—')[:10]
            apply_link = row[11] or ''
            note = str(row[16] or '—')[:20]
            status_val = str(row[15] or '待投递')
            status_e = '🟢' if status_val == '待投递' else '🔴'
            apply_cell = f'[投递]({apply_link})' if apply_link and apply_link.startswith('http') else '—'
            blines.append(f'| {company} | {job} | {location} | {deadline} | {apply_cell} | {note} | {status_e} {status_val} |')
        blines.append('')
    
    blines.append('---')
    blines.append('*每日自动更新 · 祝秋招顺利！*')
    
    with open('/workspace/docs/大公司投递表.md', 'w', encoding='utf-8') as f:
        f.write('\n'.join(blines))
    
    print(f"  Markdown已生成: {len(cats_sorted)+1} 个分类文件 + 大公司表")
    return len(cats_sorted)

# ============================================================
# Git推送
# ============================================================
def git_push():
    def run(cmd):
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True, cwd="/workspace")
        return result.stdout.strip(), result.stderr.strip()
    
    stdout, _ = run("git status --porcelain")
    if not stdout:
        print("  无变更，跳过Git提交")
        return
    
    run("git add 秋招投递追踪表.xlsx autumn_recruitment.md docs/")
    stdout, stderr = run(f'git commit -m "每日更新: {TODAY_STR} 秋招信息汇总"')
    print(f"  Git commit: {stdout}")
    
    stdout, stderr = run("git push origin main")
    if stderr and "error" in stderr.lower():
        print(f"  Git push 失败: {stderr}")
    else:
        print(f"  Git push: 成功")

# ============================================================
# 主流程
# ============================================================
def main():
    print("=" * 60)
    print(f"  秋招信息每日追踪 - {TODAY_STR}")
    print("=" * 60)
    
    # Step 1: 读取现有Excel
    print("\n[1] 读取现有Excel...")
    existing = read_existing()
    
    # Step 2: 读取Feishu数据
    print("\n[2] 读取飞书数据...")
    feishu_records = read_feishu()
    
    # Step 3: 去重
    print("\n[3] 去重处理...")
    new_records = deduplicate(existing, feishu_records)
    
    # Step 4: 更新Excel
    print("\n[4] 更新Excel...")
    new_count, total_count, all_rows = update_excel(existing, new_records)
    
    # Step 5: 更新过期状态
    print("\n[5] 更新过期状态...")
    expired_count = update_expired()
    
    # Step 6: 统计
    print("\n[6] 统计信息...")
    total, open_count, today_deadline, soon_deadline = get_stats(all_rows)
    total_soon = len(today_deadline) + len(soon_deadline)
    
    # 统计城市分布
    bj_count = sum(1 for r in new_records if '北京' in str(r.get('location', '')))
    sh_count = sum(1 for r in new_records if '上海' in str(r.get('location', '')))
    
    # 统计主要行业方向
    industries = {}
    for rec in new_records:
        ind = rec.get('industry', '其他')
        industries[ind] = industries.get(ind, 0) + 1
    top_industries = sorted(industries.items(), key=lambda x: -x[1])[:3]
    main_direction = "、".join([f"{k}({v})" for k, v in top_industries])
    
    # 统计公司
    companies = set(rec['company'] for rec in new_records)
    
    print("\n" + "=" * 60)
    print(f"  处理完成!")
    print(f"  今天新增: {new_count} 个岗位")
    print(f"  来自: {len(companies)} 家公司")
    print(f"  主要方向: {main_direction}")
    print(f"  北京: {bj_count} 个, 上海: {sh_count} 个")
    print(f"  累计: {total} 个岗位, 开放中: {open_count}")
    print(f"  7天内截止: {total_soon} 个")
    print("=" * 60)
    
    # 临近截止
    if today_deadline:
        print("\n⚠️ 今日截止岗位:")
        for r in today_deadline:
            print(f"  - {r[2]} | {r[4]} | 截止: {r[10]}")
    if soon_deadline:
        print("\n⚠️ 即将截止岗位:")
        for r, days in soon_deadline:
            print(f"  - {r[2]} | {r[4]} | {days}天后截止 ({r[10]})")
    
    # Step 7: 生成Markdown
    print("\n[7] 生成Markdown...")
    generate_markdown()
    
    # Step 8: Git推送
    print("\n[8] Git推送...")
    git_push()
    
    return new_count, len(companies), main_direction, bj_count, sh_count, total, open_count, total_soon

if __name__ == "__main__":
    main()