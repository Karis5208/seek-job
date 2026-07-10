#!/usr/bin/env python3
"""每日自动收集秋招招聘信息，整合更新到Excel和Markdown，并推送到GitHub"""

import json
import os
import re
import sys
import time
from datetime import datetime, timedelta
from collections import OrderedDict

import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ============================================================
# 配置（凭证从 .env 文件或环境变量读取）
# ============================================================
def _load_env():
    """加载 .env 文件（如果存在）"""
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.exists(env_path):
        with open(env_path, "r") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, val = line.split("=", 1)
                    os.environ.setdefault(key.strip(), val.strip())

_load_env()

FEISHU_APP_ID = os.environ.get("FEISHU_APP_ID", "")
FEISHU_APP_SECRET = os.environ.get("FEISHU_APP_SECRET", "")
FEISHU_BASE_TOKEN = os.environ.get("FEISHU_BASE_TOKEN", "")
FEISHU_TABLE_ID = os.environ.get("FEISHU_TABLE_ID", "tblO4VkZJBoMZRrV")
TODAY = datetime.now().strftime("%Y-%m-%d")
EXCEL_PATH = "/workspace/autumn_recruitment_2025.xlsx"
MD_PATH = "/workspace/autumn_recruitment.md"
DOCS_DIR = "/workspace/docs"
RAW_JSON_PATH = "/workspace/feishu_raw.json"
FILTERED_JSON_PATH = "/workspace/feishu_filtered.json"

# 金融行业关键词（该行业全部保留）
FINANCE_INDUSTRIES = [
    "金融", "证券", "银行", "保险", "投资", "基金", "信托",
    "期货", "租赁", "担保", "支付", "资产管理", "财富管理",
    "审计", "财税", "会计师事务所", "融资"
]

# 泛商科岗位关键词（非金融行业需要岗位名匹配）
BUSINESS_JOB_KEYWORDS = [
    "金融", "策略", "战略", "分析", "运营", "管培", "市场",
    "产品", "人力", "HR", "财务", "供应链", "咨询", "零售",
    "风险", "合规", "法务", "投资", "审计", "商务", "销售",
    "营销", "品牌", "公关", "客户", "研究", "经济", "数据",
    "管理", "企划", "采购", "贸易", "物流", "培训", "招聘",
    "行政", "总助", "CEO", "总裁", "助理", "秘书", "会计",
    "税务", "精算", "信评", "评级", "投行", "量化", "交易"
]

# 飞书字段映射到Excel表头
HEADERS = [
    "公司名称", "岗位名称", "岗位类别", "工作地点", "学历要求",
    "专业要求", "投递方式/链接", "开始时间", "截止时间",
    "信息来源", "收集日期", "备注", "状态"
]


# ============================================================
# 1. 飞书 API 数据获取
# ============================================================
def get_feishu_token():
    """获取飞书 tenant_access_token"""
    url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
    resp = requests.post(url, json={
        "app_id": FEISHU_APP_ID,
        "app_secret": FEISHU_APP_SECRET
    }, timeout=30)
    data = resp.json()
    if data.get("code") != 0:
        raise Exception(f"飞书认证失败: {data}")
    return data["tenant_access_token"]


def fetch_all_feishu_records(token):
    """分页获取飞书多维表格所有记录"""
    url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{FEISHU_BASE_TOKEN}/tables/{FEISHU_TABLE_ID}/records"
    all_records = []
    page_token = None
    page = 1

    while True:
        params = {"page_size": 500}
        if page_token:
            params["page_token"] = page_token

        resp = requests.get(url, headers={
            "Authorization": f"Bearer {token}"
        }, params=params, timeout=60)
        data = resp.json()

        if data.get("code") != 0:
            raise Exception(f"飞书读取失败: {data}")

        items = data.get("data", {}).get("items", [])
        all_records.extend(items)
        print(f"  第{page}页: 获取 {len(items)} 条, 累计 {len(all_records)} 条")

        has_more = data.get("data", {}).get("has_more", False)
        page_token = data.get("data", {}).get("page_token")
        if not has_more:
            break
        page += 1
        time.sleep(0.3)  # 避免触发限流

    return all_records


def is_finance_industry(industry):
    """判断是否金融行业"""
    if not industry:
        return False
    for kw in FINANCE_INDUSTRIES:
        if kw in industry:
            return True
    return False


def has_business_keyword(job_title):
    """判断岗位名是否包含泛商科关键词"""
    if not job_title:
        return False
    for kw in BUSINESS_JOB_KEYWORDS:
        if kw in job_title:
            return True
    return False


def classify_job_category(company_name, job_title, industry):
    """根据岗位名和行业分类"""
    job_title = job_title or ""
    company_name = company_name or ""

    category_rules = [
        (["金融", "银行", "证券", "保险", "投资", "基金", "信托", "期货", "融资", "信贷",
          "风控", "风险", "合规", "法务", "精算", "信评", "评级", "投行", "量化", "交易",
          "财富", "理财", "资管", "资产", "债券", "外汇", "利率", "衍生品", "股票",
          "研究员", "分析师", "行研", "固收", "权益", "宏观", "策略研究"], "金融"),
        (["战略", "策略", "企划", "规划"], "战略"),
        (["商业分析", "商分", "经营分析", "业务分析", "数据分析", "数据科学",
          "BI", "商业智能", "用户研究", "行业研究", "市场研究", "市场分析"], "商分"),
        (["运营", "内容运营", "用户运营", "活动运营", "新媒体运营", "社群运营",
          "电商运营", "平台运营", "产品运营", "增长运营"], "运营"),
        (["管培", "管培生", "培训生", "储备干部", "管理培训", "MT", "综合管理"], "管培"),
        (["市场", "营销", "品牌", "推广", "公关", "商务", "渠道", "销售",
          "客户", "广告", "策划", "增长", "BD", "GTM", "新媒体"], "市场"),
        (["产品", "产品经理", "产品设计", "产品策划", "产品助理", "PM"], "产品"),
        (["咨询", "顾问", "consulting", "advisory", "解决方案"], "咨询"),
        (["人力", "HR", "人事", "招聘", "薪酬", "培训", "组织发展", "OD",
          "员工关系", "行政", "秘书", "总助", "助理", "财务", "会计",
          "税务", "审计", "法务", "合规", "内控", "采购"], "职能"),
        (["供应链", "采购", "物流", "仓储", "SCM", "供应链管理", "贸易"], "供应链"),
        (["传媒", "媒体", "内容", "编辑", "记者", "视频", "直播", "文案",
          "设计", "UI", "UX", "视觉", "创意", "编导", "出版"], "传媒"),
        (["零售", "门店", "店铺", "陈列", "导购", "店长", "卖场"], "零售"),
    ]

    text = job_title + company_name
    for keywords, category in category_rules:
        for kw in keywords:
            if kw in text:
                return category

    return "其他"


def parse_feishu_record(record):
    """将飞书记录转为Excel行"""
    fields = record.get("fields", {})

    company = (fields.get("公司名称") or "").strip()
    job_title = (fields.get("招聘岗位") or "").strip()
    industry = (fields.get("行业类别") or "").strip()
    location = (fields.get("工作地点") or "").strip()
    deadline = (fields.get("截止日期") or "").strip()
    job_link = fields.get("岗位链接", {})
    announce_link = fields.get("招聘公告", {})
    recruit_type = (fields.get("招聘类型") or "").strip()
    recruit_year = (fields.get("招聘届别") or "").strip()
    update_time = (fields.get("更新时间") or "").strip()
    exam_info = (fields.get("笔试信息") or "").strip()
    company_size = (fields.get("公司规模") or "").strip()

    # 投递链接
    apply_link = ""
    if isinstance(job_link, dict) and job_link.get("link"):
        apply_link = job_link["link"]
    elif isinstance(announce_link, dict) and announce_link.get("link"):
        apply_link = announce_link["link"]

    # 分类
    category = classify_job_category(company, job_title, industry)

    # 备注
    notes = []
    if recruit_type:
        notes.append(f"类型:{recruit_type}")
    if recruit_year:
        notes.append(f"届别:{recruit_year}")
    if exam_info:
        notes.append(f"笔试:{exam_info}")
    if company_size:
        notes.append(f"规模:{company_size}")
    note_str = "; ".join(notes)

    return [
        company,
        job_title,
        category,
        location,
        "本科及以上",
        "金融/商科相关" if is_finance_industry(industry) else "",
        apply_link,
        update_time or "",
        deadline,
        "飞书秋招合集",
        TODAY,
        note_str,
        "开放中"
    ]


def filter_feishu_records(records):
    """筛选泛商科相关岗位"""
    filtered = []
    skipped = 0
    for record in records:
        fields = record.get("fields", {})
        industry = (fields.get("行业类别") or "").strip()
        job_title = (fields.get("招聘岗位") or "").strip()

        if is_finance_industry(industry):
            # 金融行业：全部保留
            filtered.append(record)
        elif has_business_keyword(job_title):
            # 泛商科行业：岗位名匹配才保留
            filtered.append(record)
        else:
            skipped += 1

    print(f"  筛选结果: 保留 {len(filtered)} 条, 跳过 {skipped} 条")
    return filtered


# ============================================================
# 2. 公开渠道数据采集
# ============================================================
def fetch_web_sources():
    """从公开渠道搜索当日新增招聘信息"""
    new_records = []

    # 尝试从超级简历校招汇总获取
    try:
        records = fetch_chaojijianli()
        if records:
            new_records.extend(records)
            print(f"  超级简历: 获取 {len(records)} 条")
    except Exception as e:
        print(f"  超级简历: 获取失败 - {e}")

    # 尝试从牛客网获取
    try:
        records = fetch_nowcoder()
        if records:
            new_records.extend(records)
            print(f"  牛客网: 获取 {len(records)} 条")
    except Exception as e:
        print(f"  牛客网: 获取失败 - {e}")

    return new_records


def fetch_chaojijianli():
    """从超级简历校招汇总获取招聘信息"""
    records = []
    try:
        resp = requests.get(
            "https://www.chaojijianli.com/xiaozhao/",
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            },
            timeout=15
        )
        if resp.status_code != 200:
            return records

        html = resp.text

        # 尝试提取公司名称和岗位信息
        # 匹配常见的招聘信息格式
        company_pattern = re.findall(
            r'<a[^>]*?>(?:([\u4e00-\u9fa5()A-Za-z&·]+?(?:公司|集团|银行|证券|基金|保险|信托|科技|控股|资本|咨询|会计师事务所))[^<]*?)</a>',
            html, re.IGNORECASE
        )
        # 匹配岗位关键词
        job_pattern = re.findall(
            r'([\u4e00-\u9fa5()A-Za-z/·]+?(?:管培|培训生|储备|实习生|工程师|专员|经理|分析师|顾问|运营|产品|市场|销售|财务|人力|HR|战略|咨询|研究|设计|开发|测试|算法|数据|商务|采购|供应链|风控|合规|法务|审计|会计|投资|交易|量化|精算|品牌|公关|策划|新媒体|电商|零售|物流|行政|秘书|助理|培训|招聘|教师|讲师|编辑|记者|主播|管培生))',
            html
        )

        # 提取链接
        links = re.findall(r'href=["\'](https?://[^"\']*?mp\.weixin[^"\']*?)["\']', html)
        links += re.findall(r'href=["\'](https?://[^"\']*?zhiye[^"\']*?)["\']', html)
        links += re.findall(r'href=["\'](https?://[^"\']*?hotjob[^"\']*?)["\']', html)
        links += re.findall(r'href=["\'](https?://[^"\']*?mokahr[^"\']*?)["\']', html)

        # 简单去重并构建记录
        seen_companies = set()
        for i, company in enumerate(company_pattern[:30]):
            company = company.strip()
            if company in seen_companies or len(company) < 4:
                continue
            seen_companies.add(company)

            job = job_pattern[i] if i < len(job_pattern) else "校招岗位"
            link = links[i] if i < len(links) else ""

            # 分类
            category = classify_job_category(company, job, "")
            if category == "其他":
                continue

            records.append([
                company,
                job,
                category,
                "",
                "本科及以上",
                "",
                link,
                TODAY,
                "",
                "超级简历",
                TODAY,
                "类型:校招",
                "开放中"
            ])

    except Exception as e:
        print(f"    超级简历解析异常: {e}")

    return records


def fetch_nowcoder():
    """从牛客网获取招聘信息"""
    records = []
    try:
        # 牛客网校招讨论区 API
        resp = requests.get(
            "https://www.nowcoder.com/discuss?type=7&order=1",
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            },
            timeout=15
        )
        if resp.status_code != 200:
            return records

        html = resp.text

        # 提取讨论标题(通常包含公司名+岗位)
        titles = re.findall(r'<a[^>]*?class="discuss-main[^"]*?"[^>]*?>(.*?)</a>', html, re.DOTALL)
        links = re.findall(r'<a[^>]*?class="discuss-main[^"]*?"[^>]*?href="([^"]*?)"', html)

        for i, title in enumerate(titles[:20]):
            title_clean = re.sub(r'<[^>]+>', '', title).strip()
            if not title_clean or len(title_clean) < 6:
                continue

            # 检查是否包含金融商科关键词
            if not has_business_keyword(title_clean):
                continue

            link = f"https://www.nowcoder.com{links[i]}" if i < len(links) else ""

            # 尝试从标题中提取公司名和岗位
            parts = re.split(r'[【】\[\]｜|\s]+', title_clean)
            company = parts[0] if parts else ""
            job = parts[1] if len(parts) > 1 else title_clean

            category = classify_job_category(company, job, "")
            if category == "其他":
                continue

            records.append([
                company[:30],
                job[:60],
                category,
                "",
                "本科及以上",
                "",
                link,
                TODAY,
                "",
                "牛客网",
                TODAY,
                "类型:校招讨论",
                "开放中"
            ])

    except Exception as e:
        print(f"    牛客网解析异常: {e}")

    return records


# ============================================================
# 3. 读取现有Excel数据
# ============================================================
def read_existing_data():
    """读取现有Excel中的所有数据行"""
    if not os.path.exists(EXCEL_PATH):
        return []

    wb = load_workbook(EXCEL_PATH)
    ws = wb["秋招信息汇总"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(list(row))
    wb.close()
    return rows


def mark_expired(rows):
    """标记已截止的岗位"""
    today = datetime.now()
    updated = 0
    for row in rows:
        deadline = row[8]  # 截止时间列（索引8）
        status = row[12]   # 状态列（索引12）
        if status == "已截止":
            continue
        if deadline and deadline != "招满为止":
            # 尝试解析截止日期
            try:
                dl = deadline.strip().replace(".", "-").replace("/", "-")
                dl_date = datetime.strptime(dl[:10], "%Y-%m-%d")
                if dl_date < today:
                    row[12] = "已截止"
                    updated += 1
            except ValueError:
                pass
    if updated:
        print(f"  标记已截止: {updated} 条")
    return rows


# ============================================================
# 4. 去重
# ============================================================
def deduplicate(existing_rows, new_rows):
    """按公司名称+岗位名称去重，返回新数据"""
    # 构建已有数据的key集合
    existing_keys = set()
    for row in existing_rows:
        key = (row[0] or "").strip() + "|||" + (row[1] or "").strip()[:30]
        existing_keys.add(key)

    truly_new = []
    for row in new_rows:
        key = (row[0] or "").strip() + "|||" + (row[1] or "").strip()[:30]
        if key not in existing_keys:
            truly_new.append(row)

    return truly_new


# ============================================================
# 5. 更新Excel
# ============================================================
def update_excel(existing_rows, new_rows):
    """更新Excel文件"""
    wb = load_workbook(EXCEL_PATH) if os.path.exists(EXCEL_PATH) else Workbook()

    if "秋招信息汇总" in wb.sheetnames:
        ws = wb["秋招信息汇总"]
    else:
        ws = wb.active
        ws.title = "秋招信息汇总"

    # 写表头
    for col, header in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col, value=header)

    # 排序：开放中优先，同状态内按收集日期倒序
    all_rows = existing_rows + new_rows
    # Python sort is stable, so we sort by secondary key first, then primary
    all_rows.sort(key=lambda r: str(r[10] or ""), reverse=True)  # 日期倒序
    all_rows.sort(key=lambda r: (0 if r[12] == "开放中" else 1))  # 开放中优先

    # 写入数据
    for row_idx, row_data in enumerate(all_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value or "")

    # 调整列宽
    column_widths = [18, 40, 10, 20, 10, 14, 35, 12, 12, 15, 12, 30, 10]
    for i, width in enumerate(column_widths, 1):
        col_letter = chr(64 + i) if i <= 26 else chr(64 + (i - 1) // 26) + chr(65 + (i - 1) % 26)
        ws.column_dimensions[col_letter].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(EXCEL_PATH)
    wb.close()
    print(f"  Excel已保存: {len(all_rows)} 条 (新增 {len(new_rows)} 条)")


# ============================================================
# 6. 生成Markdown
# ============================================================
def generate_markdown():
    """基于Excel数据生成Markdown文档"""
    wb = load_workbook(EXCEL_PATH)
    ws = wb["秋招信息汇总"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    today = datetime.now().strftime("%Y-%m-%d")

    # 按类别分组
    by_category = OrderedDict()
    for row in rows[1:]:
        cat = row[2] if row[2] else "其他"
        if cat not in by_category:
            by_category[cat] = []
        by_category[cat].append(row)

    # 类别排序
    cat_order = ["金融", "战略", "商分", "运营", "管培", "市场", "产品", "咨询", "职能", "供应链", "传媒", "零售"]
    cats_sorted = sorted(
        by_category.items(),
        key=lambda x: (cat_order.index(x[0]) if x[0] in cat_order else 99, -len(x[1]))
    )

    total = len(rows) - 1
    open_count = sum(1 for r in rows[1:] if r[12] == "开放中")
    closed_count = total - open_count

    # 主索引
    lines = [
        "# 2026 秋招信息汇总",
        "",
        f"> 每天早上 6:00 自动更新 | 最后更新：{today} | 共 {total} 条岗位（开放中 {open_count} | 已截止 {closed_count}）",
        "",
        "📱 **手机和网页端均可直接查看**，点击下方分类进入详情：",
        "",
        "## 岗位分类",
        "",
        "| 分类 | 数量 | 开放中 |",
        "|------|------|--------|",
    ]

    for cat, items in cats_sorted:
        safe_name = cat.replace("/", "-")
        cat_open = sum(1 for r in items if r[12] == "开放中")
        lines.append(f"| [{cat}](./docs/{safe_name}.md) | {len(items)} 条 | {cat_open} |")

    lines += [
        "",
        "---",
        "",
        "同时提供 [Excel版本](./autumn_recruitment_2025.xlsx)，下载后可筛选排序。",
        "",
        "*每日自动更新 · 祝秋招顺利！*",
    ]

    with open(MD_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    # 各分类文件
    os.makedirs(DOCS_DIR, exist_ok=True)
    for cat, items in cats_sorted:
        safe_name = cat.replace("/", "-")

        # 开放中优先
        items_sorted = sorted(items, key=lambda r: str(r[10] or ""), reverse=True)
        items_sorted = sorted(items_sorted, key=lambda r: (0 if r[12] == "开放中" else 1))

        cat_open = sum(1 for r in items_sorted if r[12] == "开放中")
        cat_closed = len(items_sorted) - cat_open

        clines = [
            f"# {cat}",
            "",
            f"> 共 {len(items)} 条 | 开放中 {cat_open} | 已截止 {cat_closed} | 更新于 {today}",
            "",
            "| 公司 | 岗位 | 地点 | 截止 | 投递 | 备注 | 状态 |",
            "|------|------|------|------|------|------|------|",
        ]

        for row in items_sorted:
            company, job, category, location, edu, major, apply, start, end, source, collect_date, note, status = row
            status_e = "🟢" if status == "开放中" else "🔴"
            company = (company or "—")[:20]
            job = (job or "—")[:35]
            location = (location or "—")[:12]
            end = (end or "—")[:10]
            note = (note or "—")[:18]
            apply_cell = f"[投递]({apply})" if apply and apply.startswith("http") else "—"
            clines.append(
                f"| {company} | {job} | {location} | {end} | {apply_cell} | {note} | {status_e} {status} |"
            )

        clines += [
            "",
            "[← 返回首页](../autumn_recruitment.md)",
        ]

        with open(f"{DOCS_DIR}/{safe_name}.md", "w", encoding="utf-8") as f:
            f.write("\n".join(clines))

    print(f"  Markdown已生成: {len(cats_sorted) + 1} 个文件")


# ============================================================
# 7. Git操作
# ============================================================
def git_commit_and_push():
    """Git提交并推送到GitHub"""
    import subprocess

    def run(cmd):
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True, cwd="/workspace")
        return result.stdout.strip(), result.stderr.strip()

    # 检查是否有变更
    stdout, _ = run("git status --porcelain")
    if not stdout:
        print("  无变更，跳过Git提交")
        return

    today = datetime.now().strftime("%Y-%m-%d")
    run("git add autumn_recruitment_2025.xlsx autumn_recruitment.md docs/ feishu_raw.json feishu_filtered.json")

    stdout, stderr = run(f'git commit -m "每日更新: {today} 秋招信息汇总"')
    print(f"  Git commit: {stdout}")

    stdout, stderr = run("git push origin main")
    if stderr and "error" in stderr.lower():
        print(f"  Git push 失败: {stderr}")
    else:
        print(f"  Git push: {stdout or '成功'}")


# ============================================================
# 主流程
# ============================================================
def main():
    print("=" * 60)
    print(f"  秋招信息每日更新 - {TODAY}")
    print("=" * 60)

    # Step 1: 读取现有Excel
    print("\n[1/6] 读取现有Excel数据...")
    existing_rows = read_existing_data()
    print(f"  现有数据: {len(existing_rows)} 条")
    existing_rows = mark_expired(existing_rows)

    # Step 2: 飞书API获取数据
    print("\n[2/6] 从飞书API获取数据...")
    try:
        token = get_feishu_token()
        print(f"  Token获取成功")
        all_records = fetch_all_feishu_records(token)
        print(f"  飞书总计: {len(all_records)} 条")

        # 保存原始数据
        with open(RAW_JSON_PATH, "w", encoding="utf-8") as f:
            json.dump(all_records, f, ensure_ascii=False, indent=2)

        # 筛选泛商科岗位
        filtered = filter_feishu_records(all_records)
        with open(FILTERED_JSON_PATH, "w", encoding="utf-8") as f:
            json.dump(filtered, f, ensure_ascii=False, indent=2)

        # 转换为Excel行
        feishu_rows = [parse_feishu_record(r) for r in filtered]
        print(f"  飞书转换: {len(feishu_rows)} 条")
    except Exception as e:
        print(f"  ⚠️ 飞书获取失败: {e}")
        feishu_rows = []

    # Step 3: 公开渠道获取
    print("\n[3/6] 从公开渠道获取招聘信息...")
    web_rows = fetch_web_sources()
    print(f"  公开渠道总计: {len(web_rows)} 条")

    # Step 4: 去重
    print("\n[4/6] 去重处理...")
    all_new = feishu_rows + web_rows
    truly_new = deduplicate(existing_rows, all_new)
    print(f"  去重后新增: {len(truly_new)} 条")

    # Step 5: 更新Excel
    print("\n[5/6] 更新Excel...")
    update_excel(existing_rows, truly_new)

    # Step 6: 生成Markdown
    print("\n[6/6] 生成Markdown...")
    generate_markdown()

    # 统计信息
    print("\n" + "=" * 60)
    print(f"  更新完成!")
    print(f"  总记录: {len(existing_rows) + len(truly_new)} 条")
    print(f"  今日新增: {len(truly_new)} 条")
    print(f"  飞书来源: {len(feishu_rows)} 条")
    print(f"  公开渠道: {len(web_rows)} 条")
    print("=" * 60)

    return truly_new


if __name__ == "__main__":
    main()